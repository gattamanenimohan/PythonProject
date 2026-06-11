from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── DATA ENTRY ────────────────────────────────────────────────────────────────
# Fill in your rows below. Each dict is one row in the matrix.
# Leave a field as "" if it has no value.

data = [
    {
        "SRS Section Number": "",
        "SRS ID": "",
        "SRS Text": "",
        "Risk Mitigating": "",
        "RMM#": "",
        "Sample Tested": "",
        "System Verification Test Procedure Covered By": "",
        "Test Evidence Document No.": "",
        "Test Statement Covered By": "",
        "Test Statement Covered By": "",
        "Test Report Reference": "",
        "Feature": "",
        "Product": "",
        "FDS Reference Document and Section (D002374620, FDS_TT_CAPIntegration.docx)": "",
    },
    # Add more rows by copying the block above ↑
]

# ── HEADERS ───────────────────────────────────────────────────────────────────
HEADERS = [
    "SRS Section Number",
    "SRS ID",
    "SRS Text",
    "Risk Mitigating",
    "RMM#",
    "Sample Tested",
    "System Verification Test Procedure Covered By",
    "Test Evidence Document No.",
    "Test Statement Covered By",
    "Test Report Reference",
    "Feature",
    "Product",
    "FDS Reference Document and Section (D002374620, FDS_TT_CAPIntegration.docx)",
]

# ── STYLES ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", start_color="1F4E79")   # dark blue
ALT_FILL      = PatternFill("solid", start_color="D6E4F0")   # light blue
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT     = Font(name="Arial", size=10)
TITLE_FONT    = Font(name="Arial", bold=True, size=14, color="1F4E79")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="B0B0B0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COL_WIDTHS = [18, 14, 35, 18, 10, 15, 38, 22, 42, 22, 18, 14, 52]

# ── BUILD WORKBOOK ────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Traceability Matrix"

# Title row
ws.merge_cells("A1:M1")
ws["A1"] = "Traceability Matrix Report"
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 30

# Header row (row 2)
for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
ws.row_dimensions[2].height = 45

# Data rows
for row_idx, row_data in enumerate(data, start=3):
    fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
        cell.font = DATA_FONT
        cell.fill = fill
        cell.alignment = LEFT_WRAP
        cell.border = BORDER
    ws.row_dimensions[row_idx].height = 40

# Column widths
for col_idx, width in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze panes (keep headers visible while scrolling)
ws.freeze_panes = "A3"

# ── SAVE ──────────────────────────────────────────────────────────────────────
OUTPUT = "/mnt/user-data/outputs/Traceability_Matrix.xlsx"
# wb.save(OUTPUT)
# print(f"Saved: {OUTPUT}")